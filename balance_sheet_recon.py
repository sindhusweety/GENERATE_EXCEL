#!/usr/bin/python3.9

#import required libraries
import datetime

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Color
from openpyxl.styles.borders import Border, Side
import locale
import pandas as pd
#openpyxl handles excelsheet
wb = Workbook()
ws = wb.active

locale.setlocale(locale.LC_ALL, 'en_US.utf8')


class Generate_Balance_Sheet:

    def __init__(self, *args):
        self.args = args
        self.excel_data = list()

    def read_excels(self):
        df = pd.read_excel('/home/sindhukumari/PycharmProjects/Upwork/Download/Open_Status_Report.XLSX')
        self.excel_data.append(df.fillna(0).values.tolist()[-1][5])
        df = pd.read_excel('/home/sindhukumari/PycharmProjects/Upwork/Download/Balance_sheet.XLSX')
        index_date = 0
        sub_res =[]
        for i in df.fillna(0).values.tolist():
            if self.args[3] in i: #example : datetime.datetime(2021, 5, 31, 0, 0)
                index_date = i.index(self.args[3])
            if self.args[2] in i: #example: 'Accounts Payable - Trade'
                sub_res.append(i)
        self.excel_data.append(sub_res[-1][index_date])

    def generate_excel(self):
        month = self.args[3].month, self.args[3].strftime("%B")
        year = self.args[3].year

        # ***********SET BACKGROUND COLOR FOR THE HEADER********************************#
        for rows in ws.iter_rows( min_col=2, max_col=6):  #min_row=1, max_row=1,
            for cell in rows:
                cell.fill = PatternFill(start_color='FFFFFF', fill_type="solid") #end_color='FFFFFF',

        # **************SET BOLD & SIZE & STYLE*************************#
        fontStyle = Font(name='Tahoma', size=15, bold=True, italic=False, vertAlign=None,
                                 underline='none', strike=False, color='FF000000')
        ws.cell(row=1, column=3, value= "Balance Sheet Reconciliation").font = fontStyle

        ws['A3'] = "Entity ID"
        ws['A4'] = "Account Number"
        ws['A5'] = "Account Description"
        ws['A6'] = "Period"

        ws['A8'] = "General Ledger Balance"
        ws['A10'] = "Supporting or Subsidiary System Detail"
        ws['A12'] = "link"

        ws['D3'] = "Prepared By"
        ws['D4'] = "Reviewed By"
        ws['F3'] = "Date"
        ws['F4'] = "Date"

        ws['E15'] = "Total Supporting or Subsidiary System Balance"
        ws['E17'] = "Variance"

        ws['E27'] = "Total of Reconciling Items"
        ws['E29'] = "Unreconciled Balance (Should be 0)"
        ws['A31'] = "Additional Explanatory Notes"

        #***************** FIX THE CELL WIDTH & HEIGHT********************#
        ws.cell(row=1, column=1).value = ""
        ws.row_dimensions[1].height = 20
        ws.cell(row=2, column=1).value = ""
        ws.column_dimensions['A'].width = 40
        ws.cell(row=2, column=2).value = ""
        ws.column_dimensions['B'].width = 30
        ws.cell(row=2, column=3).value = ""
        ws.column_dimensions['C'].width = 15
        ws.cell(row=2, column=4).value = ""
        ws.column_dimensions['D'].width = 20
        ws.cell(row=2, column=5).value = ""
        ws.column_dimensions['E'].width = 25
        ws.cell(row=2, column=6).value = ""
        ws.column_dimensions['F'].width = 20
        ws.cell(row=2, column=7).value = ""
        ws.column_dimensions['G'].width = 30

        #******************SET COLOR**********************#
        graycolor = PatternFill(start_color='D0D0D0',
                              end_color='D0D0D0',
                              fill_type='solid')
        ws['G8'].fill = graycolor

        ws['E12'].fill = graycolor
        ws['E13'].fill = graycolor
        ws['E14'].fill = graycolor

        ws['G15'].fill = graycolor
        ws['G17'].fill = graycolor

        ws['E20'].fill = graycolor
        ws['E21'].fill = graycolor
        ws['E22'].fill = graycolor
        ws['E23'].fill = graycolor
        ws['E24'].fill = graycolor
        ws['E25'].fill = graycolor

        ws['G27'].fill = graycolor
        ws['G29'].fill = graycolor

        #**************SET BOLD*************************#
        ws['B1'].font = Font(bold=True)
        ws['A8'].font = Font(bold=True)
        ws['A10'].font = Font(bold=True)
        ws['E15'].font = Font(bold=True)
        ws['E27'].font = Font(bold=True)
        ws['E29'].font = Font(bold=True)
        ws['A31'].font = Font(bold=True)

        #*************SET BORDER*********************#
        four_side_border = Border(right=Side(style='thin'),
                                  left=Side(style='thin'),
                                  top=Side(style='thin'),
                                  bottom=Side(style='thin'))
        top_bottom_header = Border(top=Side(style='thick'),
                             bottom=Side(style='thick'), diagonalUp=True, diagonalDown=True,
                 outline=True, start=None, end=None)
        for i in range(1, 8):
            ws.cell(row=1, column=i).border = top_bottom_header

        bottom_border = Border(bottom=Side(style='thick'))
        for i in range(1, 8):
            ws.cell(row=31, column=i).border = bottom_border

        '''
        Period (1-12); Month; Account Description;Account number;Entity ID name
        ACCOUNT=20100000 and ENTITYID='HWPENN' and description="Accounts Payable - Trade"
        Period (1-12); Month; Account Description;Account number;Entity ID name
        '''

        thin_border = Border(bottom=Side(style='thin'))
        ws.cell(row=3, column=2).border = thin_border
        ws.cell(row=4, column=2).border = thin_border
        ws.cell(row=5, column=2).border = thin_border
        ws.cell(row=6, column=2).border = thin_border

        ws.cell(row=3, column=5).border = thin_border
        ws.cell(row=4, column=5).border = thin_border
        ws.cell(row=3, column=7).border = thin_border
        ws.cell(row=4, column=7).border = thin_border

        ws.cell(row=8, column=7).border = thin_border
        ws.cell(row=15, column=7).border = thin_border
        ws.cell(row=17, column=7).border = thin_border
        ws.cell(row=27, column=7).border = thin_border
        ws.cell(row=29, column=7).border = thin_border

        #*********************PASS VALUES**********************#
        ws['B3'] = self.args[0]
        ws['B4'] = self.args[1]
        ws['B5'] = self.args[2]
        ws['B6'] = month[1]+'-'+str(year)
        #*************OPEN REPORT*************************************************#
        if self.excel_data[0] < 0:
            negative_amount = "("+str(locale.currency(abs(self.excel_data[0]), grouping=True))+")"
            ws['E12'] = negative_amount
            ws['E12'].font = Font(color='00FF0000')
            ws['G15'] = negative_amount
            ws['G15'].font = Font(color='00FF0000')
        else:
            ws['E12'] = locale.currency(self.excel_data[0], grouping=True) #OPEN REPORT
            ws['G15'] = locale.currency(self.excel_data[0], grouping=True) #Subsidiary

        #********************************BALANCE SHEET*************************************#
        if self.excel_data[1] < 0:
            ws['G8'] = "("+str(locale.currency(abs(self.excel_data[1]),grouping=True))+")"# Balance Sheet
            ws['G8'].font = Font(color='00FF0000')
        else:
            ws['G8'] = locale.currency(self.excel_data[1], grouping=True)

        #*********************************VARIANCE*******************************************#
        variance = self.excel_data[1] - self.excel_data[0]
        if variance < 0:
            ws['G17'] = "("+str(locale.currency(abs(variance), grouping=True)) +")" #Variance
            ws['G17'].font = Font(color='00FF0000')
        else:
            ws['G17'] = locale.currency(variance, grouping=True)

        filename = str(month[0])+'. '+month[1]+' '+str(year)+' '+self.args[1]+' - '+self.args[2]+' - '+self.args[4]+'.xlsx'
        wb.save(filename)

gObj = Generate_Balance_Sheet('****', '*****', '****', datetime.datetime(2021, 5, 31, 0, 0), '****')
gObj.read_excels()
gObj.generate_excel()





